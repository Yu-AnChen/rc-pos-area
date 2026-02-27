"""
Image processor module for validating and processing microscopy data.
"""

import os
import shutil
from pathlib import Path
from typing import List, Optional
import pandas as pd
import numpy as np
from openpyxl import load_workbook


class ValidationError(Exception):
    """Custom exception for validation errors."""
    pass


def validate_excel_file(excel_path: Path) -> List[str]:
    """
    Validate an Excel file before processing.
    
    Args:
        excel_path: Path to the Excel file
        
    Returns:
        List of error messages (empty if valid)
    """
    errors = []
    
    # Check 1: File exists
    if not excel_path.exists():
        errors.append(f"File does not exist: {excel_path}")
        return errors  # Can't continue without file
    
    # Check 2: Valid Excel format with required sheets
    try:
        xls = pd.ExcelFile(excel_path)
        sheet_names = xls.sheet_names
        sheet_names_lower = [s.lower() for s in sheet_names]
        
        if 'files' not in sheet_names_lower:
            errors.append("Missing 'Files' sheet")
        if 'thresholds' not in sheet_names_lower:
            errors.append("Missing 'Thresholds' sheet")
        
        if errors:
            return errors  # Can't continue without required sheets
        
        # Get actual sheet names (case-insensitive)
        files_sheet_idx = sheet_names_lower.index('files')
        thresholds_sheet_idx = sheet_names_lower.index('thresholds')
        files_sheet_name = sheet_names[files_sheet_idx]
        thresholds_sheet_name = sheet_names[thresholds_sheet_idx]
        
        # Check 3: Files sheet has exactly one row
        files_df = pd.read_excel(excel_path, sheet_name=files_sheet_name)
        if len(files_df) == 0:
            errors.append("'Files' sheet is empty")
            return errors
        elif len(files_df) > 1:
            errors.append(f"'Files' sheet has {len(files_df)} rows, expected 1")
            return errors
        
        # Check 4: Files sheet has required columns
        if 'Slide Name' not in files_df.columns:
            errors.append("'Files' sheet missing 'Slide Name' column")
        if 'File Path' not in files_df.columns:
            errors.append("'Files' sheet missing 'File Path' column")
        
        if errors:
            return errors
        
        # Check 5: Image file exists
        image_path = files_df.loc[0, 'File Path']
        if pd.isna(image_path):
            errors.append("Image file path is empty")
            return errors
        
        image_path = Path(image_path)
        if not image_path.exists():
            errors.append(f"Image file does not exist: {image_path}")
            return errors
        
        # Check 6: palom.reader.OmePyramidReader can be initiated
        try:
            import palom
            reader = palom.reader.OmePyramidReader(str(image_path))
            
            # Check 7: Channel numbers are valid
            thresholds_df = pd.read_excel(excel_path, sheet_name=thresholds_sheet_name)
            
            if 'Channel #' not in thresholds_df.columns:
                errors.append("'Thresholds' sheet missing 'Channel #' column")
                return errors
            
            if 'Threshold' not in thresholds_df.columns:
                errors.append("'Thresholds' sheet missing 'Threshold' column")
                return errors
            
            # Check that channel 2 exists (required for tissue region mask)
            if 2 not in thresholds_df['Channel #'].values:
                errors.append("Channel #2 is required for tissue region definition but not found in 'Thresholds' sheet")
                return errors
            
            # Determine pyramid level and get channel count
            # Using same pyramid level as in original script
            PYRAMID_LEVEL = 2
            
            try:
                num_channels = len(reader.pyramid[PYRAMID_LEVEL])
            except (IndexError, KeyError):
                errors.append(f"Image does not have pyramid level {PYRAMID_LEVEL}")
                return errors
            
            # Check channel numbers (Excel uses 1-based indexing)
            for idx, row in thresholds_df.iterrows():
                channel_num = row['Channel #']
                if pd.isna(channel_num):
                    errors.append(f"Row {idx+2} in 'Thresholds' sheet has empty Channel #")
                elif not isinstance(channel_num, (int, np.integer)):
                    errors.append(f"Row {idx+2} in 'Thresholds' sheet has non-integer Channel # ({channel_num})")
                elif channel_num < 1 or channel_num > num_channels:
                    errors.append(f"Channel # {channel_num} is invalid (image has {num_channels} channels: 1-{num_channels})")
            
        except ImportError:
            errors.append("Cannot import 'palom' library - is it installed?")
        except Exception as e:
            errors.append(f"Cannot read image file: {e}")
        
    except Exception as e:
        errors.append(f"Cannot read Excel file: {e}")
    
    return errors


def process_single_excel(excel_path: Path, output_dir: Path, 
                         verbose: bool = False, quiet: bool = False) -> Path:
    """
    Process a single Excel file.
    
    Args:
        excel_path: Path to input Excel file
        output_dir: Path to output directory
        verbose: Print detailed progress
        quiet: Minimal output
        
    Returns:
        Path to output file
    """
    import palom
    import dask_image.ndfilters
    
    # Constants (same as original script)
    PYRAMID_LEVEL = 2
    GAUSSIAN_SIGMA = 1
    PIXEL_SIZE = 0.325 * 2**PYRAMID_LEVEL
    
    # Create output directory
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Check if output directory is writable
    test_file = output_dir / '.test_write'
    try:
        test_file.touch()
        test_file.unlink()
    except Exception as e:
        raise ValidationError(f"Cannot write to output directory {output_dir}: {e}")
    
    # Generate output filename
    output_filename = f"{excel_path.stem}_processed.xlsx"
    output_path = output_dir / output_filename
    
    if verbose:
        print(f"   Reading {excel_path.name}...")
    
    # Load Excel file
    xls = pd.ExcelFile(excel_path)
    sheet_names = xls.sheet_names
    sheet_names_lower = [s.lower() for s in sheet_names]
    
    files_sheet_idx = sheet_names_lower.index('files')
    thresholds_sheet_idx = sheet_names_lower.index('thresholds')
    files_sheet_name = sheet_names[files_sheet_idx]
    thresholds_sheet_name = sheet_names[thresholds_sheet_idx]
    
    files_df = pd.read_excel(excel_path, sheet_name=files_sheet_name)
    thresholds_df = pd.read_excel(excel_path, sheet_name=thresholds_sheet_name)
    
    # Get image file path
    slide_name = files_df.loc[0, 'Slide Name']
    file_path = files_df.loc[0, 'File Path']
    
    if verbose:
        print(f"   Loading image: {Path(file_path).name}")
    
    # Load image
    reader = palom.reader.OmePyramidReader(str(file_path))
    
    # Set Channel # as index
    thresholds_df = thresholds_df.set_index('Channel #')
    
    # Process each channel
    if verbose:
        print(f"   Processing {len(thresholds_df)} channel(s)...")
    
    # First, create tissue region mask from channel 2
    if verbose:
        print(f"   Creating tissue region mask from channel 2...")
    
    channel_2_threshold = thresholds_df.loc[2, 'Threshold']
    channel_2_idx = 2 - 1  # Convert to 0-based
    tissue_img = reader.pyramid[PYRAMID_LEVEL][channel_2_idx].astype('float32')
    tissue_gimg = dask_image.ndfilters.gaussian_filter(tissue_img, GAUSSIAN_SIGMA)
    tissue_mask = tissue_gimg > channel_2_threshold
    tissue_area = int(np.sum(tissue_mask)) * PIXEL_SIZE**2
    
    if verbose:
        print(f"   Tissue area: {tissue_area:.2f} µm^2")
    
    # Now process all channels (including channel 2)
    for cc, tt in list(thresholds_df['Threshold'].items()):
        # Convert from 1-based (Excel) to 0-based (Python) indexing
        channel_idx = cc - 1
        img = reader.pyramid[PYRAMID_LEVEL][channel_idx].astype('float32')
        gimg = dask_image.ndfilters.gaussian_filter(img, GAUSSIAN_SIGMA)
        
        # Overall metrics
        thresholds_df.loc[cc, 'Area (µm^2)'] = gimg.size * PIXEL_SIZE**2
        thresholds_df.loc[cc, 'Positive Area (µm^2)'] = int(np.sum(gimg > tt)) * PIXEL_SIZE**2
        
        # Tissue region metrics
        thresholds_df.loc[cc, 'Tissue Area (µm^2)'] = tissue_area
        positive_in_tissue = np.sum((gimg > tt) & tissue_mask)
        thresholds_df.loc[cc, 'Positive Area in Tissue (µm^2)'] = int(positive_in_tissue) * PIXEL_SIZE**2
    
    # Calculate positive fractions
    thresholds_df['Positive Fraction (%)'] = 100 * (
        thresholds_df['Positive Area (µm^2)'] / thresholds_df['Area (µm^2)']
    )
    thresholds_df['Positive Fraction in Tissue (%)'] = 100 * (
        thresholds_df['Positive Area in Tissue (µm^2)'] / thresholds_df['Tissue Area (µm^2)']
    )
    
    # Round values
    area_cols = ['Area (µm^2)', 'Positive Area (µm^2)', 'Tissue Area (µm^2)', 'Positive Area in Tissue (µm^2)']
    thresholds_df[area_cols] = thresholds_df[area_cols].round(2)
    thresholds_df['Positive Fraction (%)'] = thresholds_df['Positive Fraction (%)'].round(5)
    thresholds_df['Positive Fraction in Tissue (%)'] = thresholds_df['Positive Fraction in Tissue (%)'].round(5)
    
    if verbose:
        print(f"   Writing results to {output_filename}...")
    
    # Copy original Excel file to output
    shutil.copy2(excel_path, output_path)
    
    # Load workbook and update Thresholds sheet
    wb = load_workbook(output_path)
    
    # Remove old Thresholds sheet if exists
    if thresholds_sheet_name in wb.sheetnames:
        del wb[thresholds_sheet_name]
    
    # Create new sheet and write data
    ws = wb.create_sheet(thresholds_sheet_name)
    
    # Reset index to write Channel # as a column
    thresholds_df_out = thresholds_df.reset_index()
    
    # Write header
    for c_idx, col in enumerate(thresholds_df_out.columns, 1):
        ws.cell(row=1, column=c_idx, value=col)
    
    # Write data
    for r_idx, row in enumerate(thresholds_df_out.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Move Thresholds sheet to second position (after Files)
    wb.move_sheet(thresholds_sheet_name, offset=-(len(wb.sheetnames) - 2))
    
    wb.save(output_path)
    
    return output_path
