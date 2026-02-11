"""
Report generator module for creating summary reports from processed files.
"""

from pathlib import Path
from typing import List, Dict, Tuple
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
import matplotlib.pyplot as plt


# Tab10 color palette (RGB tuples)
TAB10_COLORS = [
    (31, 119, 180),   # blue
    (255, 127, 14),   # orange
    (44, 160, 44),    # green
    (214, 39, 40),    # red
    (148, 103, 189),  # purple
    (140, 86, 75),    # brown
    (227, 119, 194),  # pink
    (127, 127, 127),  # gray
    (188, 189, 34),   # olive
    (23, 190, 207),   # cyan
]


def rgb_to_hex(rgb: Tuple[int, int, int]) -> str:
    """Convert RGB tuple to hex color string."""
    return '{:02X}{:02X}{:02X}'.format(*rgb)


def get_channel_signature(thresholds_df: pd.DataFrame) -> Tuple[int, ...]:
    """
    Get the channel signature (sorted tuple of channel numbers) from a thresholds dataframe.
    
    Args:
        thresholds_df: DataFrame with 'Channel #' column
        
    Returns:
        Sorted tuple of channel numbers
    """
    channels = sorted(thresholds_df['Channel #'].unique())
    return tuple(channels)


def assign_groups(processed_files: List[Path], verbose: bool = False) -> Dict[Tuple[int, ...], List[Dict]]:
    """
    Group processed files by their channel configuration.
    
    Args:
        processed_files: List of paths to processed Excel files
        verbose: Print detailed progress
        
    Returns:
        Dictionary mapping channel signatures to list of file info dicts
    """
    groups = {}
    
    for file_path in processed_files:
        try:
            # Read Files and Thresholds sheets
            xls = pd.ExcelFile(file_path)
            sheet_names_lower = [s.lower() for s in xls.sheet_names]
            
            if 'files' not in sheet_names_lower or 'thresholds' not in sheet_names_lower:
                if verbose:
                    print(f"   Skipping {file_path.name}: missing required sheets")
                continue
            
            files_idx = sheet_names_lower.index('files')
            thresholds_idx = sheet_names_lower.index('thresholds')
            
            files_df = pd.read_excel(file_path, sheet_name=xls.sheet_names[files_idx])
            thresholds_df = pd.read_excel(file_path, sheet_name=xls.sheet_names[thresholds_idx])
            
            # Get slide name
            slide_name = files_df.loc[0, 'Slide Name']
            
            # Get channel signature
            channel_sig = get_channel_signature(thresholds_df)
            
            # Add to groups
            if channel_sig not in groups:
                groups[channel_sig] = []
            
            groups[channel_sig].append({
                'file_path': file_path,
                'slide_name': slide_name,
                'files_df': files_df,
                'thresholds_df': thresholds_df,
            })
            
        except Exception as e:
            if verbose:
                print(f"   Error reading {file_path.name}: {e}")
            continue
    
    # Sort files within each group by slide name
    for channel_sig in groups:
        groups[channel_sig].sort(key=lambda x: x['slide_name'])
    
    return groups


def generate_summary_report(processed_files: List[Path], output_path: Path,
                           verbose: bool = False, quiet: bool = False):
    """
    Generate a summary report from processed Excel files.
    
    Args:
        processed_files: List of paths to processed Excel files
        output_path: Path for output summary file
        verbose: Print detailed progress
        quiet: Minimal output
    """
    if not quiet:
        print(f"\nGenerating summary report...")
    
    # Group files by channel configuration
    if verbose:
        print(f"   Grouping files by channel configuration...")
    
    groups = assign_groups(processed_files, verbose)
    
    if not groups:
        raise ValueError("No valid processed files found")
    
    # Assign group numbers and colors
    group_assignments = {}
    sorted_channel_sigs = sorted(groups.keys())
    
    for group_num, channel_sig in enumerate(sorted_channel_sigs, 1):
        color_idx = (group_num - 1) % len(TAB10_COLORS)
        group_assignments[channel_sig] = {
            'group_num': group_num,
            'color': TAB10_COLORS[color_idx],
            'hex_color': rgb_to_hex(TAB10_COLORS[color_idx]),
        }
    
    if not quiet:
        print(f"   Found {len(groups)} group(s):")
        for channel_sig, info in group_assignments.items():
            print(f"      Group {info['group_num']}: Channels {list(channel_sig)} ({len(groups[channel_sig])} file(s))")
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create summary sheet
    if verbose:
        print(f"   Creating summary sheet...")
    
    summary_ws = wb.create_sheet('Summary', 0)
    
    # Summary sheet headers
    summary_headers = ['Slide Name', 'File Path', 'Group']
    for c_idx, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=1, column=c_idx, value=header)
        cell.font = Font(bold=True)
    
    # Add data to summary sheet
    row_idx = 2
    for channel_sig in sorted_channel_sigs:
        group_info = group_assignments[channel_sig]
        group_num = group_info['group_num']
        hex_color = group_info['hex_color']
        
        for file_info in groups[channel_sig]:
            # Write slide name and file path
            summary_ws.cell(row=row_idx, column=1, value=file_info['slide_name'])
            summary_ws.cell(row=row_idx, column=2, value=str(file_info['file_path']))
            summary_ws.cell(row=row_idx, column=3, value=f"Group {group_num}")
            
            # Apply color to row
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
            for c_idx in range(1, 4):
                summary_ws.cell(row=row_idx, column=c_idx).fill = fill
            
            row_idx += 1
    
    # Auto-size columns in summary sheet
    for column in summary_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create individual sheets for each slide
    if verbose:
        print(f"   Creating individual slide sheets...")
    
    for channel_sig in sorted_channel_sigs:
        group_info = group_assignments[channel_sig]
        hex_color = group_info['hex_color']
        
        for file_info in groups[channel_sig]:
            slide_name = file_info['slide_name']
            thresholds_df = file_info['thresholds_df']
            
            # Create sheet with slide name
            # Excel sheet names must be <= 31 chars
            sheet_name = slide_name[:31] if len(slide_name) > 31 else slide_name
            
            # Handle duplicate sheet names
            original_name = sheet_name
            counter = 1
            while sheet_name in wb.sheetnames:
                suffix = f"_{counter}"
                max_base_len = 31 - len(suffix)
                sheet_name = original_name[:max_base_len] + suffix
                counter += 1
            
            ws = wb.create_sheet(sheet_name)
            
            # Set sheet tab color
            ws.sheet_properties.tabColor = hex_color
            
            # Write thresholds data
            # Headers
            for c_idx, col in enumerate(thresholds_df.columns, 1):
                cell = ws.cell(row=1, column=c_idx, value=col)
                cell.font = Font(bold=True)
            
            # Data
            for r_idx, row in enumerate(thresholds_df.itertuples(index=False), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    if verbose:
        print(f"   Saving to {output_path.name}...")
    
    wb.save(output_path)
    
    if not quiet:
        print(f"\nReport summary:")
        print(f"   Total files: {sum(len(files) for files in groups.values())}")
        print(f"   Groups: {len(groups)}")
        print(f"   Sheets created: {len(wb.sheetnames)} (1 summary + {len(wb.sheetnames)-1} slides)")
